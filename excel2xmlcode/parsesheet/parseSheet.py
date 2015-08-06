__author__ = 'Administrator'

from openpyxl.utils import *
from excel2xmlcode import readConfig
from excel2xmlcode.parsesheet import parseSheetSimple
from excel2xmlcode.parsesheet import parseSheetMultipleTags
from excel2xmlcode.parsesheet import parseSheetMultipleFiles

class LanguageSheet:
    def readSheet(self, tws):
        self.lsws=tws  #对应的表格
        self.files=[]  #存储  [.xml文件名,ETTree数据]
        #解析表格
        self.parseSheet()
    #解析表格
    def parseSheet(self):
        lstree,lstrees=None,None
        for row in self.lsws.iter_rows():
            for cell in row:
                if(cell.value==readConfig.keyColumn):#读取到  cell内容为 Key的关键字，表格开始解析
                    if(column_index_from_string(cell.column)==1):#cell.column=A， cell.row=1
                        #第一种类型的表格，表格的名字对应一个.xml，表格中只有key值，无需指定tag值
                        lstree=parseSheetSimple.parseSimpleSheet(self.lsws,cell.row+1,self.lsws.max_row,column_index_from_string(cell.column),self.lsws.max_column,column_index_from_string(cell.column),cell.row,None)
                        self.files.append([self.lsws.title,lstree])
                    elif(column_index_from_string(cell.column)==2 and self.lsws.cell(row=cell.row, column=1).value==readConfig.tagColumn):
                        #第二种类型表格，表格的名字对应一个.xml，表格中有key值，tag值
                        lstree=parseSheetMultipleTags.parseMultipleTagsSheet(self.lsws,cell.row+1,self.lsws.max_row,column_index_from_string(cell.column),self.lsws.max_column,column_index_from_string(cell.column),cell.row)
                        self.files.append([self.lsws.title,lstree])
                    elif(column_index_from_string(cell.column)==2 and self.lsws.cell(row=cell.row,column=1).value==readConfig.fileColumn):
                        #第三种类型表格，表格的名字无意义，表格中含有导出的.xml文件名，tag值
                        lstrees=parseSheetMultipleFiles.parseMultipleFilessSheet(self.lsws,cell.row+1,self.lsws.max_row,column_index_from_string(cell.column),self.lsws.max_column,column_index_from_string(cell.column),cell.row)
                        self.files.extend(lstrees)
                    return

