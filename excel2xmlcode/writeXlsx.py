__author__ = 'Administrator'
from openpyxl import Workbook
def writeSimpleXlsx():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws['A1']=42
    ws.append([1,2,3])
    #ws['A2']=10 #datetime.datetime.now()
    ws1 = wb.create_sheet()
    ws1.title="sheet2"
    wb.save("sample.xlsx")